# -*- coding: utf-8 -*-


from selenium.webdriver.common.action_chains import ActionChains  
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from webdriver_manager.chrome import ChromeDriverManager 
from selenium.common.exceptions import NoSuchElementException

import requests
from bs4 import BeautifulSoup
import time

import os, sys
from openpyxl import Workbook
from openpyxl import load_workbook

# достает html код по указанной ссылке
def get_html(url):       
    print('Gettin HTML-code from ', url)
    driver.get(url)
    try:
        elem = driver.find_element_by_class_name("vacancy-serp")
        returningstring = elem.get_attribute('innerHTML')
        return returningstring
    except NoSuchElementException:
        return ''

 # достает html код по указанной ссылке
def get_html_contact_info(url):        
    driver.get(url)
    try:        
        elem = driver.find_element_by_class_name("vacancy-contacts")
        elem1 = elem.find_element_by_class_name("bloko-link-switch") 
        if(elem1):       
            elem1.click()
        elem = elem.get_attribute('innerHTML')
        fio=''
        phone=''
        email=''
        elem = BeautifulSoup(str(elem), 'lxml') 
        try: 
            fio = elem.find_all(attrs={"data-qa": "vacancy-contacts__fio"})
            if(fio[0]):
              fio = fio[0].getText() 
        except NoSuchElementException:
            fio=''
        try:   
            phones = []          
            phone = elem.find_all(attrs={"data-qa": "vacancy-contacts__phone"})  
            for ph in phone:
                phones.append(ph.getText())

        except NoSuchElementException:
            phone=''
        try: 
            email = elem.find_all(attrs={"data-qa": "vacancy-contacts__email"})
            if(len(email)):
               email = email[0].getText() 
        except NoSuchElementException:
            email=''       
        returningstring = []
        returningstring.append(fio)
        returningstring.append(phones)
        returningstring.append(email)
        return returningstring
    except NoSuchElementException:
        return ''       

def search_data(elem, string):
    try: 
      s = elem.find_element_by_xpath(string)
      returningstring = s.text
      return returningstring
    except NoSuchElementException:
        return ''

# проверяет, есть ли на странице ссылки на вакансии
def is_empty(html):
    soup = BeautifulSoup(html, 'lxml')
    links = soup.find_all('a', class_='HH-LinkModifier')
    if links == []:
        return True
    else:
        return False


# функция, которая для данного запроса и региона ищет все страницы с результатами поиска и набирает большой список со всеми ссылками на вакансии
# возвращает список ссылок по запросу query в регионе с кодом area
def get_all_offers_links(query, area):
    # headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
    url_base = 'https://hh.ru/search/vacancy'
    url_text = '?text='+query
    url_area = '&area='+area
    url_page = '&page='

    # когда не найдем с помощью bs4 нужный элемент, то выставим его False
    # нужен для остановки цикла перебора всех страниц
    page_is_not_empty = True

    all_links = []
    page = 0

    while page_is_not_empty:
        url = url_base + url_text + url_area + url_page + str(page)
        time.sleep(.5)
        html = get_html(url)
        if not is_empty(html):
            all_links = get_offers_links(html, all_links)
            page += 1
        else:
            page_is_not_empty = False

    return all_links


# функция, которая собирает все ссылки на вакансии на странице поиска
# принимает список, который уже может быть не пустой, возвращает дополненный список
def get_offers_links(html, all_links):
    # новый объект класса BeutifulSoup
    soup = BeautifulSoup(html, 'lxml')
 
    links = soup.find_all('div', class_='vacancy-serp-item')
    for link in links:
        vacancy_done=[]
        vacancy = BeautifulSoup(str(link), 'lxml')
        vacancy_link=vacancy.find('a', class_='HH-LinkModifier') 
        vacancy_href = vacancy_link.get('href').split('?') #Взяли ссылку на вакансию      
        vacancy_text = vacancy_link.string #Взяли текст вакансии
        vacancy_done.append(vacancy_href[0]) 
        vacancy_done.append(vacancy_text)
        vacancy_link_company = vacancy.find('a', class_='bloko-link_secondary')
        if(vacancy_link_company):
            vacancy_href_company = vacancy_link_company.get('href').split('?') #Взяли ссылку на компанию
            vacancy_text_company = vacancy_link_company.string #Взяли ссылку на компанию
        else:
            vacancy_href_company[0] = ''
            vacancy_text_company = ''
        vacancy_done.append("https://hh.ru"+vacancy_href_company[0])
        vacancy_done.append(vacancy_text_company)
        vacancy_city_company=vacancy.find('span', class_='vacancy-serp-item__meta-info')
        vacancy_city_company=vacancy_city_company.string
        vacancy_done.append(vacancy_city_company)
        all_links.append(vacancy_done)
    return all_links

def parse_offers(links, query): 

    for link in links:
       html = get_html_contact_info(link[0])
       link.append(html)
       time.sleep(.3)

    wb = load_workbook('companyes.xlsx') 
    if not query in wb.sheetnames: 
        ws = wb.create_sheet(query) # insert at the end (default)  
    ws['A1'] = 'Компания'
    ws['B1'] = 'Ссылка на компанию'
    ws['C1'] = 'Название вакансии'
    ws['D1'] = 'Ссылка на вакансию'
    ws['E1'] = 'Город'
    ws['F1'] = 'Контактная информация:Имя'
    ws['G1'] = 'Контактная информация:Телефоны'
    ws['H1'] = 'Контактная информация:E-mail'
 
    row=2
    for link in links:
        ws.cell(row=row, column=1, value=str(link[3])) 
        ws.cell(row=row, column=2).hyperlink = str(link[2])
        ws.cell(row=row, column=3, value=str(link[1]))   
        ws.cell(row=row, column=4).hyperlink = str(link[0])
        ws.cell(row=row, column=5, value=str(link[4]))
        if(link[5]):
            ws.cell(row=row, column=6, value=str(link[5][0]))
        if(link[5]): 
            for l in link[5][1]:
                str5=''
                if(ws.cell(row=row, column=7).value):
                    str5 = ws.cell(row=row, column=7).value
                str5=str(str5)+" "+str(l)    
                ws.cell(row=row, column=7, value=str(str5))
        if(link[5]):         
            ws.cell(row=row, column=8, value=str(link[5][2]))     
        row=row+1 
    wb.save('companyes.xlsx')
        
 
def begin(queryes):
    for query in queryes:
        time.sleep(1)
         # сначала вытащим все ссылки на вакансии по данному запросу и региону
        links = get_all_offers_links(query, area)
        # теперь распарсим информацию по каждой ссылке, полученной выше
        parse_offers(links, query)

        print('Проверено ', len(links), ' вакансий.')
 

 

if __name__ == '__main__': 
    options = webdriver.ChromeOptions()
    options.add_argument("--enable-javascript")
    driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)

    #query = 'solidworks'
    query = ['php','python','javascript','js','mysql']
    area = '113' 
    
    begin(query)

 
 




 